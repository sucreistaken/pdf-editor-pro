"""
PDF to Excel Tool - PDF tablolarini Excel'e donustur
"""
import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


def _try_convert_number(value):
    """Sayi gibi gorunen string'leri sayiya donustur"""
    if not isinstance(value, str) or not value.strip():
        return value
    stripped = value.strip()
    # Turk/Avrupa formati: 1.234,56
    if re.match(r'^-?\d{1,3}(\.\d{3})*(,\d+)?$', stripped):
        cleaned = stripped.replace('.', '').replace(',', '.')
        try:
            return float(cleaned) if '.' in cleaned else int(cleaned)
        except (ValueError, TypeError):
            return value
    # Ingiliz formati: 1,234.56
    if re.match(r'^-?\d{1,3}(,\d{3})*(\.\d+)?$', stripped):
        cleaned = stripped.replace(',', '')
        try:
            return float(cleaned) if '.' in cleaned else int(cleaned)
        except (ValueError, TypeError):
            return value
    # Basit sayi: 123 veya 123.45
    if re.match(r'^-?\d+(\.\d+)?$', stripped):
        try:
            return float(stripped) if '.' in stripped else int(stripped)
        except (ValueError, TypeError):
            return value
    return value


def pdf_to_excel(input_path, output_path):
    """
    PDF icerisindeki tablolari Excel dosyasina donustur

    Args:
        input_path (str): Girdi PDF yolu
        output_path (str): Cikti XLSX yolu

    Returns:
        dict: Islem sonucu
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sayfa 1"

        total_tables = 0
        total_rows = 0
        current_row = 1

        with pdfplumber.open(input_path) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                tables = page.extract_tables()

                if not tables:
                    # Tablo yoksa metni cikar
                    text = page.extract_text()
                    if text:
                        if page_idx > 0:
                            ws = wb.create_sheet(title=f"Sayfa {page_idx + 1}")
                            current_row = 1
                        for line in text.split('\n'):
                            line = line.strip()
                            if not line:
                                continue
                            # 2+ bosluk ile split yaparak sutunlara dagit
                            parts = re.split(r'\s{2,}', line)
                            if len(parts) > 1:
                                for col_idx, part in enumerate(parts):
                                    cell_value = _try_convert_number(part.strip())
                                    ws.cell(row=current_row, column=col_idx + 1, value=cell_value)
                            else:
                                ws.cell(row=current_row, column=1, value=line)
                            current_row += 1
                            total_rows += 1
                    continue

                for table_idx, table in enumerate(tables):
                    total_tables += 1

                    if page_idx > 0 or table_idx > 0:
                        sheet_name = f"Sayfa {page_idx + 1}"
                        if table_idx > 0:
                            sheet_name += f" - Tablo {table_idx + 1}"
                        ws = wb.create_sheet(title=sheet_name[:31])
                        current_row = 1

                    for row in table:
                        if row is None:
                            continue
                        for col_idx, cell in enumerate(row):
                            value = cell if cell is not None else ''
                            value = _try_convert_number(value)
                            ws.cell(row=current_row, column=col_idx + 1, value=value)
                        current_row += 1
                        total_rows += 1

        # Sutun genisliklerini otomatik ayarla
        for sheet in wb.worksheets:
            for col_cells in sheet.columns:
                max_length = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[col_letter].width = max(adjusted_width, 8)

        wb.save(output_path)

        return {
            'success': True,
            'message': f'{total_tables} tablo, {total_rows} satır çıkarıldı',
            'tables_found': total_tables,
            'total_rows': total_rows,
            'page_count': len(wb.sheetnames)
        }

    except Exception as e:
        logger.error(f"pdf_to_excel error: {e}", exc_info=True)
        return {'success': False, 'error': 'İşlem başarısız'}
